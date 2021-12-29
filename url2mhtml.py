#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Dec 20 09:24:45 2021

@author: shalo
"""
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from concurrent.futures import ThreadPoolExecutor,ProcessPoolExecutor

import re, os
import hashlib
import xlrd
import openpyxl
import time


def validateTitle(title):
    rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/ \ : * ? " < > |'
    mstr = r"[\'“”\s']"
    new_title = re.sub(rstr, "_", title)  # 替换为下划线
    new_title = re.sub(mstr, "", title) #删除特殊字符
    return new_title


def get_profile():
    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument('--headless')  # 谷歌无头模式
    chromeOptions.add_argument('--disable-gpu')  # 禁用显卡
    chromeOptions.add_argument('window-size=1280,800')  # 指定浏览器分辨率
    chromeOptions.add_argument("--no-sandbox")
    return chromeOptions


def get_browser():
    #with webdriver.Chrome(ChromeDriverManager().install()) as browser:
    #    return browser
    browser = webdriver.Chrome(options=get_profile())
    return browser

def _get_page(initial_url,save_path):    
    
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')
    
    #with webdriver.Chrome(ChromeDriverManager().install()) as driver:   
    with webdriver.Chrome(options=options) as driver: 
        #driver.add_argument('--disable-gpu')
        driver.get(initial_url)
        # ページをMHTML形式で保存
        time.sleep(3)
        mhl = driver.execute_cdp_cmd("Page.captureSnapshot", {})
        
        if mhl:
            pass
        else:#未获取到内容，等待6秒后再试一次
            time.sleep(6)
            mhl = driver.execute_cdp_cmd("Page.captureSnapshot", {})
            if mhl:
                pass
            else:#未获取到内容，等待12秒后再试一次
                time.sleep(12)
                mhl = driver.execute_cdp_cmd("Page.captureSnapshot", {})
            
        #time.sleep(12)
        mtitle = driver.title
        #print(mtitle)
        mtitle = validateTitle(mtitle)
        #print(mtitle)

        #存储为mhtml
        get_time = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())
        filename = os.path.join(save_path, '{}.mhtml'.format(mtitle+get_time))
        print(filename)
        with open(filename, 'w', newline='') as f:
            f.write(mhl['data'])
        
        driver.close()
    
def snapshot_page(url_file,result_path):
    input_xlsx= openpyxl.load_workbook(url_file)
    #sheet_names=input_xlsx.get_sheet_names()
    ws = input_xlsx[r"输入"]
    
    # 获取sheet的最大行数和列数
    rows = ws.max_row
    #cols = ws.max_column

    for i in range(1,rows):   
        country  =  ws.cell(i+1,1).value # 国家地区       
        web = ws.cell(i+1,2).value #网站名称
        Section = ws.cell(i+1,3).value #板块名称
        url = ws.cell(i+1,4).value   #url地址
        getdate = time.strftime("%Y-%m-%d", time.localtime()) #当前日期
        #print(country)
        #print(web)
      
        #创建首级目录-国家
        newpath = os.path.join(result_path,country)        
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        
        #创建次级目录-网站
        newpath = os.path.join(newpath,web)        
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            
        #以板块创建子目录
        newpath = os.path.join(newpath,Section)        
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            
        #以当前日期创建子目录
        newpath = os.path.join(newpath,getdate) 
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            
        try:
            _get_page(url,newpath)
        except :
            print("save mhtml error:")
            print(url)
            #print(newpath)
            #time.sleep(5)
            pass
            #break


if __name__ == '__main__':
    #page_url = f"https://qiita.com/mochi_yu2/items/e2480ae3b2a6db9d7a98"
    save_path = r"/home/shalo/downloadMhtml"
    url_file = r"./config/urls5.xlsx"
    
    snapshot_page(url_file,save_path)
    #os.path.exists(save_path)
    #_get_page(page_url,save_path)